import os
import cv2

import json
import pytube
import pickle
import pandas
import base64
import zipfile
import openpyxl
from pathlib import Path
import ipywidgets as widgets
from google.colab import files
import rich
from rich.markdown import Markdown
from moviepy.editor import VideoFileClip
from manager.util import Fairy, Cipher, WebMaster
from manager.video import Video as MyVideo
from IPython.display import clear_output
from flask import Flask, jsonify, request, send_from_directory
from IPython.core.magic import register_line_magic, register_cell_magic


fairy = Fairy()
cipher = Cipher()
my_video = MyVideo(video_path='videos/dumbbell legs.mp4')
web_master = None

################################
# widgets
################################

wid_login_file = widgets.FileUpload(
    description='Upload login file', accept='.login',  multiple=False)
wid_import_btn = widgets.Button(
    description='Import login file', disabled=False)

wid_account = widgets.Password(description='Account')
wid_password = widgets.Password(description='Password')
wid_username = widgets.Dropdown(options=list(
    fairy.data_config.keys()), description='Username')
wid_project = widgets.Text(description='Project')

wid_login_btn = widgets.Button(description='Login', disabled=False)


wid_upload_videos_btn = widgets.Button(
    description='Upload Videos', disabled=False)
wid_upload_datas_btn = widgets.Button(
    description='Upload Datas', disabled=False)
wid_refresh_btn = widgets.Button(
    description='Refresh Videos and Datas', disabled=False)

wid_video = widgets.Dropdown(
    options=[(item.name, item)
             for i, item in enumerate(list(fairy.video_root_path.glob('*.mp4')) +
                                      list(fairy.video_output_root_path.glob('*.mp4')))
             ],
    description='Video : ',
)
wid_data = widgets.Dropdown(
    options=[(item.name, item)
             for i, item in enumerate(list(fairy.data_root_path.glob('*.*')) +
                                      list(fairy.data_output_root_path.glob('*.*')))
             ],
    description='Data : ',
)


def import_btn(wid_import_btn):
    _key = list(wid_login_file.value.keys())[0]
    _content = eval(base64.b64decode(
        wid_login_file.value[_key]['content']).decode('UTF-8'))
    wid_account.value = _content['account']
    wid_password.value = _content['password']
    wid_username.value = _content['username']
    wid_project.value = _content['project']


def login_btn(wid_login_btn):
    global web_master 
    cipher.u8(wid_account.value, wid_password.value)
    if wid_username.value in fairy.data_config.keys():
        if cipher.check(fairy.data_config[wid_username.value]):
            _user = wid_username.value
            data_cipher = fairy.data_config[_user]
            _check_list = ['ngrok_t', 'git_m', 'repo_p', 'cont_p'] if _user == 'master' else ['ngrok', 'git', 'repo_p', 'cont_p']
            web_master = WebMaster(cipher=cipher, user=_user, check_list=_check_list, check_dict=data_cipher)
            msg_pass = """# Setup complete, welcome to use the action label system"""
            _md = Markdown(msg_pass)
            rich.console.Console().print(_md)
    else:
        msg_wrong = """
        # Welcome The Johnson Action Label System

        您好請登入再使用喔

        1. 你可能帳密輸打錯或使用者選錯
        2. 如果你還沒註冊請去 [註冊](https://www.google.com/)
        """
        _md = Markdown(msg_wrong)
        rich.console.Console().print(_md)
        


def focus_video_fun():
    return list(fairy.video_root_path.glob('*.mp4')) \
        + list(fairy.video_output_root_path.glob('*.mp4'))


def focus_data_fun():
    return list(fairy.data_root_path.glob('*.xlsx')) \
        + list(fairy.data_output_root_path.glob('*.xlsx')) \
        + list(fairy.data_root_path.glob('*.json')) \
        + list(fairy.data_output_root_path.glob('*.json')) \
        + list(fairy.data_root_path.glob('*.pickle')) \
        + list(fairy.data_output_root_path.glob('*.pickle'))


def upload_videos_btn(wid_upload_videos_btn):
    uploaded = files.upload()
    for filename, file_data in uploaded.items():
        os.rename(filename, (fairy.video_root_path/filename))
    # update the videos dropdown
    wid_video.options = [(item.name, item)
                         for i, item in enumerate(focus_video_fun())]


def upload_datas_btn(wid_upload_datas_btn):
    uploaded = files.upload()
    for filename, file_data in uploaded.items():
        os.rename(filename, (fairy.data_root_path/filename))
    # update the datas dropdown
    wid_data.options = [(item.name, item)
                        for i, item in enumerate(focus_data_fun())]


def refresh_btn(wid_refresh_btn):
    wid_video.options = [(item.name, item)
                         for i, item in enumerate(focus_video_fun())]
    wid_data.options = [(item.name, item)
                        for i, item in enumerate(focus_data_fun())]


wid_import_btn.on_click(import_btn)
wid_login_btn.on_click(login_btn)
wid_upload_videos_btn.on_click(upload_videos_btn)
wid_upload_datas_btn.on_click(upload_datas_btn)
wid_refresh_btn.on_click(refresh_btn)


wid_download_btn = widgets.Button(
    description='Download', disabled=False)

wid_download = widgets.Dropdown(
    options=[('All', 0),
             ('videos', 1),
             ('output datas', 2),
             ('output videos', 3),
             ('output datas and videos', 4),],
    description='Download Folders: ',
)


def download_btn(wid_download_btn):
    if wid_download.value == 0:
        folders_to_compress = [str(fairy.data_root_path),
                               str(fairy.data_output_root_path),
                               str(fairy.video_root_path),
                               str(fairy.video_output_root_path),]
    elif wid_download.value == 1:
        folders_to_compress = [str(fairy.video_root_path),]
    elif wid_download.value == 2:
        folders_to_compress = [str(fairy.data_output_root_path),]
    elif wid_download.value == 3:
        folders_to_compress = [str(fairy.video_output_root_path),]
    elif wid_download.value == 4:
        folders_to_compress = [str(fairy.data_output_root_path),
                               str(fairy.video_output_root_path),]

    zip_filename = fairy.temp_root_path / 'compressed_folders.zip'
    with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for folder in folders_to_compress:
            for root, dirs, _files in os.walk(folder):
                for _file in _files:
                    file_path = os.path.join(root, _file)
                    relative_path = os.path.relpath(file_path, folder)
                    zipf.write(file_path, arcname=relative_path)

    files.download(zip_filename)


wid_download_btn.on_click(download_btn)


################################
# decorate function
################################


def cipher_check(function):
    def wrapper(*args, **kwargs):
        clear_output()
        if cipher._flag:
            return function(*args, **kwargs)
        else:
            rich.print('請登入以後再使用')
    return wrapper


################################
# tools
################################


@register_line_magic
def login_tool(line):
    clear_output()
    display(
        widgets.VBox(
            [
                widgets.HBox([wid_login_file, wid_import_btn]),
                widgets.HBox([wid_account, wid_password,
                             wid_username, wid_project]),
                widgets.HBox([wid_login_btn]),
            ]
        )
    )

# quick display show
@register_line_magic
def path_display(line): 
    return display(widgets.HBox([wid_video, wid_refresh_btn])) if eval(line) else display(widgets.HBox([wid_data, wid_refresh_btn]))


@cipher_check
def path_manager(line):
    display(
        widgets.VBox(
            [
                widgets.HBox(
                    [wid_video, wid_upload_videos_btn, wid_refresh_btn]),
                widgets.HBox([wid_data, wid_upload_datas_btn]),
            ]
        )
    )


@cipher_check
def download_tool(line):
    display(
        widgets.HBox([wid_download, wid_download_btn])
    )


@cipher_check
def video_show_tool(line):
    if len(line.split(' ')) == 1:
        _rate = "50%" if line == "" else (line + "%")
        max_duration = 45
    else:
        _line = line.split(' ')
        _rate = _line[0] + "%"
        max_duration = min(int(_line[1]), 80)

    cap = cv2.VideoCapture(str(wid_video.value))
    duration = cap.get(7)/cap.get(5)
    if duration > max_duration:
        display(VideoFileClip(str(wid_video.value)).subclip(
            0, max_duration).ipython_display(maxduration=(max_duration+1), width=_rate))
    else:
        display(VideoFileClip(str(wid_video.value)).ipython_display(
            maxduration=(max_duration+1), width=_rate))


@cipher_check
def data_show_tool(line):
    _sub = wid_data.value.name.rsplit('.', 1)[1]
    if _sub == 'xlsx':
        display(pandas.read_excel(wid_data.value))
    elif _sub == 'json':
        with open(wid_data.value, 'r') as f:
            rich.print(json.load(f))
    elif _sub == 'pickle':
        with open(wid_data.value, 'r') as f:
            rich.print(pickle.load(f))
    else:
        rich.print('not supported yet!')


@cipher_check
def demo_data_prepare(line):
    pyt_obj = pytube.YouTube(fairy.demo_video_url)
    pyt_obj.streams.filter().get_highest_resolution().download(filename='demo_video.mp4')
    os.rename('demo_video.mp4', (fairy.video_root_path/'demo_video.mp4'))
    wid_video.options = [(item.name, item)
                         for i, item in enumerate(focus_video_fun())]
    wid_data.value = Path('datas/demo_content.xlsx')
    wid_video.value = Path('videos/demo_video.mp4')


@cipher_check
def excel_cut_video(line):
    # basic class label
    sample_class_json = {
        "version": "v1.2.1",
        "video_id": "",
        "brightcove_url": "",
        "actions_label": []
    }

    wb = openpyxl.load_workbook(wid_data.value)
    sheet = wb.active
    action_number = 0
    for i in range(sheet.max_row):
        if sheet.cell(i+1, 4).value and sheet.cell(i+1, 4).value != '-' and i+1 > 3:
            action_number += 1

            start = sheet.cell(i+1, 7).value
            start_time = start.hour*60*60 + start.minute*60 + start.second
            end = sheet.cell(i+1, 8).value
            end_time = end.hour*60*60 + end.minute*60 + end.second
            duration = end_time - start_time

            sample_class_json["actions_label"].append({
                "workout_name": sheet.cell(i+1, 3).value,
                "set": int(sheet.cell(i+1, 5).value),
                "repeats": int(sheet.cell(i+1, 6).value),
                "start_time": start_time,
                "end_time": end_time,
                "duration": duration,
                "action_label": {}
            })
    print(f"Excel 偵測到 {action_number} 個動作，開始剪輯影片")
    video = VideoFileClip(str(wid_video.value))

    for count, action in enumerate(sample_class_json['actions_label']):
        if count < int(line):
            print(f"Cutting : {count}/{action_number}")
            output = video.subclip(action['start_time'], action['end_time'])
            output.write_videofile(f"{str(fairy.video_root_path)}/{wid_project.value}_video_{count+1}.mp4",
                                   temp_audiofile="temp/temp-audio.m4a", remove_temp=True, codec="libx264", audio_codec="aac", verbose=False)

    with open(fairy.data_output_root_path / (wid_project.value + '_class.json'), 'w') as f:
        json.dump(sample_class_json, f)

    wid_video.options = [(item.name, item)
                         for i, item in enumerate(focus_video_fun())]
    print('All Done')


@cipher_check
def video_cut_fit(line):
    _time = line.split(' ')
    video = VideoFileClip(str(wid_video.value)).subclip(
        int(_time[0]), int(_time[1]))
    video.write_videofile(f"{str(fairy.video_output_root_path)}/{wid_video.value.name.rsplit('.',1)[0]+'_fit.mp4'}",
                          temp_audiofile="temp/temp-audio.m4a", remove_temp=True, codec="libx264", audio_codec="aac", verbose=False)
    wid_video.options = [(item.name, item)
                         for i, item in enumerate(focus_video_fun())]


@cipher_check
def video_preprocessing(line):
    global my_video
    my_video = MyVideo(video_path=str(wid_video.value), frame_index=0)
    my_video.annotate_square_one()


@cipher_check
def video_preprocessing_classical(line):
    global my_video
    _var = line.split(' ')
    if _var[0] == 'New':     
        display(my_video.draw_rectangle(0, 0, 0, True, False))
    else:
        my_video = MyVideo(video_path=str(wid_video.value), frame_index=int(_var[4]))
        display(my_video.draw_rectangle(int(_var[1]), int(_var[2]), int(_var[3]), False, False))



@cipher_check
def video_cut_square(line):
    my_video.video_cutting(flip_flag=bool(line), show=False)


@cipher_check
def model_inference_video(line):
    fairy.model_inference_video()


@cipher_check
def modify_pickle(line):
    _var = line.split(' ')
    fairy.modify_pickle(frame_index=int(_var[1]),
                        body_index=int(_var[2]),
                        x_rate=float(_var[3]),
                        y_rate=float(_var[4]),
                        save_pickle_check=eval(_var[0]),)


@cipher_check
def json_scheme_check(line):
    fairy.json_scheme_check(sample_json=eval(line))


@cipher_check
def write_json_file(line):
    fairy.write_json_file(sample_json=eval(line))


@cipher_check
def simplify_json_file(line):
    fairy.simplify_json_file(json_path=str(wid_data.value))


@cipher_check
def output_hash_function(line):
    fairy.hash_function(json_path=str(wid_data.value), remove_version_check=eval(line))


@cipher_check
def flask_api_server(line):
    _var = line.split(' ')
    url_method = _var[0]
    url_path = _var[1]

    app = Flask(__name__)
    web_master.run_with_ngrok(app)

    @app.route('/project_name', methods = ['GET'])
    def proj_name():
        if request.method == 'GET':
            return wid_project.value
    
    @app.route(url_path + '/<string:value>', methods = [url_method])
    def sent_json(value):
        with open(wid_data.value, 'r') as f:
            data = json.load(f)
        if request.method == url_method:
            return jsonify(data)

    @app.route('/get_video/<string:path>')
    def get_file(path):
        return send_from_directory('videos_output', 'target.mp4', as_attachment=True)

    app.run()



################################
# register line magic
################################

def register_tools(line):
    fun_list = line.split(' ')
    for fun in fun_list:
        template = f"@register_line_magic \ndef {fun}_run(line):\n    {fun}(line)"
        exec(template)


line_string = ' '.join([
    'path_manager',
    'download_tool',
    'video_show_tool',
    'data_show_tool',
    'demo_data_prepare',
    'excel_cut_video',
    'video_cut_fit',
    'video_preprocessing',
    'video_preprocessing_classical',
    'video_cut_square',
    'model_inference_video',
    'modify_pickle',
    'json_scheme_check',
    'write_json_file',
    'simplify_json_file',
    'output_hash_function',
    'flask_api_server',
])


################################
# register line runner
################################
register_tools(line_string)
